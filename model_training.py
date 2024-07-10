# -*- coding: utf-8 -*-
"""
Created on Mon Jul  8 18:03:31 2024

@author: TimOConnor
"""

"""
Code to run the model training. Takes the input config file, trains the 
appropriate model and outputs the results.

Assumes all data is in the raw_data folder within the main_dir defined by
model_configuration.py. Within all_data, expects subfolders 'test' and 'train' 
each with subsequent subfolders 'Benign' and 'Malignant'.

Each class folder expected to contain 224x224 jpeg images of skin leasions. 
This code will train a classifier to classify between 'Benign' and 'Malignant'
skin leasions

This code will output the metrics of accurcay, sensitivity, specificity, auc
and mcc to an excel spreadsheet. 
If the trained model is the best performing model, the model and its outputs
 (training curves, ROC curve, confusion matrix, results.pkl) will be saved
to the model_results folder within the working directory. 

"""
import os
import numpy as np
import cv2
import matplotlib.pyplot as plt
import pandas as pd
import pickle
import random

import openpyxl
from openpyxl.styles import PatternFill

import torch
import torch.nn as nn
import torch.optim as optim
from torchvision import models
from torch.utils.data import DataLoader, Dataset
from sklearn.metrics import recall_score, matthews_corrcoef, accuracy_score, roc_curve, roc_auc_score
from sklearn.metrics import confusion_matrix
from sklearn.metrics import ConfusionMatrixDisplay
from utils import AddGaussNoise, get_simple_CNN
from model_configuration import get_config
from sklearn.model_selection import train_test_split

import albumentations as A
import albumentations.pytorch


# Define dataloader class to load in the images 
class SkinLesionDataset(Dataset):
    def __init__(self, image_paths, labels, transform=None):
        self.image_paths = image_paths
        self.labels = labels
        self.transform = transform

    def __len__(self):
        return len(self.image_paths)

    def __getitem__(self, idx):
        img_path = self.image_paths[idx]
        image = cv2.imread(img_path)
        image = cv2.cvtColor(image, cv2.COLOR_BGR2RGB)  # Convert BGR to RGB
        image = image.astype(np.float32) / 255.0  # Set all data between 0 and 1
        label = self.labels[idx]

        if self.transform:
            image = self.transform(image=image)["image"]
        return image, label, img_path



# Define augmentation and preprocessing pipelines
# use random rotations of the images, small shifts and scales, additive gaussian noise and pixel drouput. 
# Normalize using imagenet normalization constants snce we start with a pretrained ResNet
data_transforms = {
    'train': A.Compose([
        A.RandomRotate90(p=1),   
        A.ShiftScaleRotate(shift_limit=0.15, scale_limit = .05, rotate_limit=45, p=.2),
        AddGaussNoise(mean_val=0, var_val=.001,p=0.2),
        A.PixelDropout (dropout_prob=0.1, per_channel=True, drop_value=0, mask_drop_value=None, always_apply=False, p=.2),
        A.Normalize(mean=(0.485, 0.456, 0.406), std=(0.229, 0.224, 0.225)),
        A.pytorch.ToTensorV2()
    ]),
    'test': A.Compose([
        A.Normalize(mean=(0.485, 0.456, 0.406), std=(0.229, 0.224, 0.225)),
        A.pytorch.ToTensorV2()
    ])
} 

# Function to train a CNN model with early stopping and return the best model
def train_model(model, device, criterion, optimizer, dataloaders, dataset_sizes, num_epochs=25, patience=10):
    best_model_wts = model.state_dict()
    best_loss = float('inf')
    best_acc = 0.0
    epochs_no_improve = 0

    train_loss_history = []
    val_loss_history = []
    train_acc_history = []
    val_acc_history = []

    for epoch in range(num_epochs):
        print(f'Epoch {epoch}/{num_epochs - 1}')
        print('-' * 10)

        for phase in ['train', 'val']:
            if phase == 'train':
                model.train()  # Set model to training mode
            else:
                model.eval()   # Set model to evaluate mode

            running_loss = 0.0
            running_corrects = 0
            

            for inputs, labels, _ in dataloaders[phase]:
                inputs = inputs.to(device)
                labels = labels.to(device)

                optimizer.zero_grad()

                with torch.set_grad_enabled(phase == 'train'):
                    outputs = model(inputs)
                    _, preds = torch.max(outputs, 1)
                    loss = criterion(outputs, labels)

                    if phase == 'train':
                        loss.backward()
                        optimizer.step()

                running_loss += loss.item() * inputs.size(0)
                running_corrects += torch.sum(preds == labels.data)

            epoch_loss = running_loss / dataset_sizes[phase]
            epoch_acc = running_corrects.double() / dataset_sizes[phase]

            if phase == 'train':
                train_loss_history.append(epoch_loss)
                train_acc_history.append(epoch_acc.cpu().item())
            else:
                val_loss_history.append(epoch_loss)
                val_acc_history.append(epoch_acc.cpu().item())

            print(f'{phase} Loss: {epoch_loss:.4f} Acc: {epoch_acc:.4f}')

            # Deep copy the model if validation loss has decreased
            if phase == 'val' and epoch_loss < best_loss:
                best_loss = epoch_loss
                best_acc = epoch_acc
                best_model_wts = model.state_dict()
                epochs_no_improve = 0
            elif phase == 'val':
                epochs_no_improve += 1

        # Early stopping
        if epochs_no_improve >= patience:
            print(f'Early stopping triggered after {epoch} epochs.')
            break

    # Load best model weights
    model.load_state_dict(best_model_wts)

    # Plot training and validation loss and accuracy
    plt.figure(figsize=(12, 4))
    plt.subplot(1, 2, 1)
    plt.plot(train_loss_history, label='Train Loss')
    plt.plot(val_loss_history, label='Val Loss')
    plt.xlabel('Epochs')
    plt.ylabel(' Loss')
    plt.legend()
    plt.title('Training and Validation Loss Curves')

    plt.subplot(1, 2, 2)
    plt.plot(train_acc_history, label='Train Acc')
    plt.plot(val_acc_history, label='Val Acc')
    plt.xlabel('Epochs')
    plt.ylabel('Accuracy')
    plt.legend()
    plt.title('Training and Validation Classification Accuracy')

    plt.show()
    
    training_curves = {'Train_loss': train_loss_history,'Train_acc': train_acc_history,'Val_loss': val_loss_history,'Val_acc': val_acc_history}

    return model, best_loss, best_acc, training_curves

def train_lesion_detector(config):
    
    torch.manual_seed(config['seed'])
    np.random.seed(config['seed'])
    
    # Set device
    device = torch.device("cuda:0" if torch.cuda.is_available() else "cpu")
    
    
    train_root_dir = 'raw_data/train'
    test_root_dir = 'raw_data/test'
    
    # Initialize lists to store paths and labels
    train_image_paths = []
    train_labels = []
    test_image_paths = []
    test_labels = []
    
    # Load training data paths and labels
    for label, subdir in enumerate(['Benign', 'Malignant']):
        subdir_path = os.path.join(os.getcwd(),train_root_dir, subdir)
        for img_name in os.listdir(subdir_path):
            img_path = os.path.join(subdir_path, img_name)
            train_image_paths.append(img_path)
            train_labels.append(label)
    
    # Load test data paths and labels
    for label, subdir in enumerate(['Benign', 'Malignant']):
        subdir_path = os.path.join(os.getcwd(),test_root_dir, subdir)
        for img_name in os.listdir(subdir_path):
            img_path = os.path.join(subdir_path, img_name)
            test_image_paths.append(img_path)
            test_labels.append(label)
        
    # Split the training data so we have a validation set to optimize the model training
    train_image_paths, val_image_paths, train_labels, val_labels = train_test_split(
    train_image_paths, train_labels, test_size=config['val_frac'], random_state=42, stratify=train_labels)
    
    #Use only a subset of the trianing data to speed up training during prototyping
    if config['train_frac'] != 1:
        num_train_samples = int(len(train_image_paths) * config['train_frac'])
        train_image_paths = random.sample(train_image_paths, num_train_samples)

    train_dataset = SkinLesionDataset(train_image_paths, train_labels, transform=data_transforms['train'])
    val_dataset = SkinLesionDataset(val_image_paths, val_labels, transform=data_transforms['test'])
    test_dataset = SkinLesionDataset(test_image_paths, test_labels,transform=data_transforms['test']) 

    # Create dataloaders
    train_loader = DataLoader(train_dataset, batch_size=config['batch_size'], shuffle=True)
    val_loader = DataLoader(val_dataset, batch_size=config['batch_size'], shuffle=False)
    test_loader = DataLoader(test_dataset, batch_size=config['batch_size'], shuffle=False)

    dataloaders = {'train': train_loader, 'val': val_loader, 'test': test_loader}
    dataset_sizes = {'train': len(train_dataset), 'val': len(val_dataset), 'test': len(test_dataset)}

    # Load the model to train
    if config['use_simple_CNN']:
        model = get_simple_CNN()
        
    else:
        # Use ResNet by default, start with ImageNet weights
        model = models.resnet18(weights='DEFAULT')

        # Freeze layers except the final fully connected layer to speed up training
        for param in model.parameters():
            param.requires_grad = False
    
        # Unfreeze the final two layers to increase trainable capacity
        for name, param in model.named_parameters():
            if "layer4" in name or "layer3" in name:
                param.requires_grad = True
    
        num_ftrs = model.fc.in_features
        model.fc = nn.Linear(num_ftrs, len(config['classes']))
    
    
    model = model.to(device)

    criterion = nn.CrossEntropyLoss()
    optimizer = optim.Adam(model.parameters(), lr=config['initial_lr'])

    model, best_loss, best_acc, training_curves = train_model(model, device, criterion, optimizer, dataloaders, dataset_sizes, num_epochs=config['max_epochs'], patience=config['early_stop_patience'])


    print(f"Training completed with best validation loss: {best_loss:.4f} and best validation accuracy: {best_acc:.4f}")

    # Evaluate the model on the test set
    model.eval()
    all_labels = []
    all_preds = []
    all_probs = []
    fp_paths = []
    fn_paths = []

    with torch.no_grad():
        for inputs, labels, paths in dataloaders['test']:
            inputs = inputs.to(device)
            labels = labels.to(device)
            outputs = model(inputs)
            _, preds = torch.max(outputs, 1)
            probs = nn.functional.softmax(outputs, dim=1)[:, 1]
            all_labels.extend(labels.cpu().numpy())
            all_preds.extend(preds.cpu().numpy())
            all_probs.extend(probs.cpu().numpy())

            for i in range(len(labels)):
                if preds[i] == 1 and labels[i] == 0:
                    fp_paths.append(paths[i])
                elif preds[i] == 0 and labels[i] == 1:
                    fn_paths.append(paths[i])

    confMat = confusion_matrix(all_labels, all_preds)
    disp = ConfusionMatrixDisplay(confusion_matrix=confMat, display_labels=config['classes'])
    disp.plot(cmap=plt.cm.Blues, colorbar=False)
    plt.title("Classification confusion matrix (Test data)")

    # Calculate metrics
    accuracy = accuracy_score(all_labels, all_preds)
    sensitivity = recall_score(all_labels, all_preds)
    tn, fp, fn, tp = confusion_matrix(all_labels, all_preds).ravel()
    specificity = tn / (tn + fp)
    mcc = matthews_corrcoef(all_labels, all_preds)
    auc = roc_auc_score(all_labels, all_probs)

    print(f"Accuracy: {accuracy:.4f}")
    print(f"Sensitivity (Recall): {sensitivity:.4f}")
    print(f"Specificity: {specificity:.4f}")
    print(f"MCC: {mcc:.4f}")
    print(f"AUC: {auc:.4f}")

    # Plot ROC curve
    fpr, tpr, _ = roc_curve(all_labels, all_probs)
    plt.figure(figsize=(8, 6))
    plt.plot(fpr, tpr, color='blue', lw=2, label=f'ROC curve (AUC = {auc:.2f})')
    plt.plot([0, 1], [0, 1], color='gray', linestyle='--')
    plt.xlim([0.0, 1.0])
    plt.ylim([0.0, 1.05])
    plt.xlabel('False Positive Rate')
    plt.ylabel('True Positive Rate')
    plt.title('Receiver Operating Characteristic (ROC) Curve')
    plt.legend(loc='lower right')
    plt.show()

    
    results = {'Acc': accuracy, 'Sens': sensitivity, 'Spec': specificity, 'MCC': mcc, 'AUC': auc, 'TP':tp, 'TN': tn, 'FP': fp, 'FN': fn, 
            'FPs': list(fp_paths), 'FNs': list(fn_paths), 'Labels': all_labels, 'Preds': all_preds, 'Probs': all_probs, 'Training_curves': training_curves}
    
    # Define the path to the Excel file
    save_dir = os.path.join(os.getcwd(),'model_results')
    excel_file = os.path.join(save_dir,'model_results.xlsx')

    # Save the model results
    results_to_save = {
        'Model': [config['model_name']],
        'Accuracy': [results['Acc']],
        'Sensitivity': [results['Sens']],
        'Specificity': [results['Spec']],
        'MCC': [results['MCC']],
        'AUC': [results['AUC']]
    }
    
    
    # Check if the Excel file exists
    if os.path.exists(excel_file):
        # Load existing data from Excel
        existing_df = pd.read_excel(excel_file)
        if config['model_name'] in existing_df['Model'].values:
            new_name = config['model_name']+'_v'
            vers = existing_df['Model'].str.contains(new_name).sum()+1
            new_name = new_name+str(vers)
            config['model_name'] = new_name
            results_to_save['Model'] =  config['model_name']
        results_df = pd.DataFrame(results_to_save)
        updated_df = pd.concat([existing_df, results_df], ignore_index=True)
    else:
        results_df = pd.DataFrame(results_to_save)
        updated_df = results_df 
        
    updated_df.to_excel(excel_file, index=False)
    print(f"Results saved to {excel_file}")
    
    # Determine if this is the best model so far, if so, save the model and the plots
    max_accuracy_model = updated_df.loc[updated_df['Accuracy'].idxmax(), 'Model']
    # Highlight the row with the highest accuracy using green fill
    workbook = openpyxl.load_workbook(excel_file)
    sheet = workbook.active
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=1):
        for cell in row:
            if cell.value == max_accuracy_model:
                for col in range(1, sheet.max_column + 1):
                    sheet.cell(row=cell.row, column=col).fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
    workbook.save(excel_file)


    if max_accuracy_model == config['model_name']:
        model_name = config['model_name']+'.pth'
        torch.save(model, os.path.join(save_dir,model_name))
        # Save config dictionary to a text file
        with open(os.path.join(save_dir, 'best_config.txt'), 'w') as f:
            for key, value in config.items():
                f.write(f'{key}: {value}\n')
        
        # Save the results
        results_file = os.path.join(save_dir, 'best_model_results.pkl')
        with open(results_file, 'wb') as f:
            pickle.dump(results, f)
        
        # Save the plots
        # ROC Curve
        fpr, tpr, _ = roc_curve(results['Labels'], results['Probs'])
        plt.figure(figsize=(8, 6))
        plt.plot(fpr, tpr, color='blue', lw=2, label=f'ROC curve (AUC = {auc:.2f})')
        plt.plot([0, 1], [0, 1], color='gray', linestyle='--')
        plt.xlim([0.0, 1.0])
        plt.ylim([0.0, 1.05])
        plt.xlabel('False Positive Rate')
        plt.ylabel('True Positive Rate')
        plt.title('Receiver Operating Characteristic (ROC) Curve')
        plt.legend(loc='lower right')
        plt.savefig(os.path.join(save_dir,'ROC_curve.png'))
        plt.close()
        
        # Training Curves
        plt.figure(figsize=(12, 4))
        plt.subplot(1, 2, 1)
        plt.plot(results['Training_curves']['Train_loss'], label='Train Loss')
        plt.plot(results['Training_curves']['Val_loss'], label='Val Loss')
        plt.xlabel('Epochs')
        plt.ylabel(' Loss')
        plt.legend()
        plt.title('Training and Validation Loss Curves')

        plt.subplot(1, 2, 2)
        plt.plot(results['Training_curves']['Train_acc'], label='Train Acc')
        plt.plot(results['Training_curves']['Val_acc'], label='Val Acc')
        plt.xlabel('Epochs')
        plt.ylabel('Accuracy')
        plt.legend()
        plt.title('Training and Validation Classification Accuracy')
        plt.savefig(os.path.join(save_dir,'Training_curves.png'))
        plt.close()
        
        # Confusion Matrix
        confMat = confusion_matrix(results['Labels'], results['Preds'])
        disp = ConfusionMatrixDisplay(confusion_matrix=confMat, display_labels=config['classes'])
        disp.plot(cmap=plt.cm.Blues, colorbar=False)
        plt.title("Classification confusion matrix (Test data)")
        plt.savefig(os.path.join(save_dir,'Confusion_matrix.png'))
        plt.close()
        

    return results

if __name__ == '__main__':
    config = get_config()
    results = train_lesion_detector(config)





























